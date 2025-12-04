import streamlit as st
import pandas as pd
import extract_msg
import io
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="AI4IAM - Special Permissions", page_icon="📧", layout="wide")

# --- HEADER ---
st.title("📧 Special Permissions Review ")
st.markdown("""
<div style="font-size:0.8rem; color:gray; margin-top:-10px;">
Created by Loïc Persyn with the help of AI
</div>
""", unsafe_allow_html=True)

st.markdown("---")


# --- FUNCTION TO EXTRACT INFO FROM EMAIL BODY ---
def extract_info(email_body):
    info = {
        'User_ID': None,
        'User_Name': None,
        'Manager': None,
        'Special_Permission': None,
        'Permission_Code': None,
        'End_Date': None,
        'Link': None
    }

    text = email_body.replace("\r", "").strip()

    match_user = re.search(r'User:\s*([A-Z0-9]+)\s*/\s*([^\r\n]+)', text, re.IGNORECASE)
    if match_user:
        info['User_ID'] = match_user.group(1).strip()
        info['User_Name'] = match_user.group(2).strip()

    match_manager = re.search(r'Manager:\s*([^\r\n]+)', text, re.IGNORECASE)
    if match_manager:
        info['Manager'] = match_manager.group(1).strip()

    match_permission = re.search(
        r'Special\s*permission[:\-]?\s*(.+)\((\d{3,})\)',
        text,
        re.IGNORECASE
    )
    if match_permission:
        info['Special_Permission'] = match_permission.group(1).strip()
        info['Permission_Code'] = match_permission.group(2).strip()

    match_date = re.search(r'Planned End date:\s*(\d{2}[-/]\d{2}[-/]\d{4})', text, re.IGNORECASE)
    if match_date:
        info['End_Date'] = match_date.group(1).strip()

    match_link = re.search(r'Link:\s*<?([^>\r\n]+)>?', text, re.IGNORECASE)
    if match_link:
        link_text = match_link.group(1).strip()
        url_match = re.search(r'(https?://[^\s>]+)', link_text)
        if url_match:
            link_text = url_match.group(1).strip()
        info['Link'] = link_text

    return info


# --- HTML TABLE GENERATOR FOR .EML ---
def df_to_html_table(df):
    html = '<table style="border-collapse: collapse; width: 100%;">'
    html += '<thead><tr>'

    yellow_col = 'Needs Extension ? [y/n]'

    # Header
    for col in df.columns:
        bg = 'background-color: yellow;' if col == yellow_col else ''
        html += f'<th style="border:1px solid black; padding:4px; font-weight:bold; {bg}">{col}</th>'
    html += '</tr></thead><tbody>'

    # Rows
    for _, row in df.iterrows():
        html += '<tr>'
        for col in df.columns:
            bg = 'background-color: yellow;' if col == yellow_col else ''
            value = row[col] if pd.notna(row[col]) else ''
            if col == "Link" and isinstance(value, str) and value.startswith("http"):
                value = f'<a href="{value}">{value}</a>'
            html += f'<td style="border:1px solid black; padding:4px; {bg}">{value}</td>'
        html += '</tr>'
    html += '</tbody></table>'
    return html


# --- SIDEBAR FILE UPLOAD ---
st.sidebar.image("AI4IAM Logo (2).png", width=240)
st.sidebar.header("📁 Upload Files")
uploaded_files = st.sidebar.file_uploader(
    "Select your .msg files",
    type=['msg'],
    accept_multiple_files=True
)


# --- MAIN LOGIC ---
if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} file(s) uploaded")

    if st.button("🚀 EXTRACT DATA", type="primary"):
        data = []
        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, uploaded_file in enumerate(uploaded_files):
            try:
                msg = extract_msg.Message(io.BytesIO(uploaded_file.read()))
                info = extract_info(msg.body)
                data.append(info)
                msg.close()

                progress_bar.progress((i + 1) / len(uploaded_files))
                status_text.text(f"Processing: {i+1}/{len(uploaded_files)}")
            except Exception as e:
                st.warning(f"⚠️ Error with {uploaded_file.name}: {str(e)}")

        if data:

            # --- CREATE SORTED DATAFRAME ---
            df = pd.DataFrame(data)
            df.insert(6, 'Needs Extension ? [y/n]', "")

            col_order = [
                'User_ID', 'User_Name', 'Manager',
                'Special_Permission', 'Permission_Code',
                'End_Date', 'Needs Extension ? [y/n]', 'Link'
            ]
            df = df[col_order]

            df.sort_values(by=['Manager', 'User_Name'], inplace=True)

            # --- DISPLAY ---
            st.subheader("📊 Extracted Data (Sorted)")
            st.dataframe(df, use_container_width=True)

            # --- EXCEL EXPORT ---
            df_excel = df.copy()
            df_excel['Link'] = df_excel['Link'].apply(
                lambda x: f'=HYPERLINK("{x}", "{x}")' if isinstance(x, str) and x.startswith("http") else x
            )

            output_excel = io.BytesIO()
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                df_excel.to_excel(writer, index=False, sheet_name="Permissions")
                ws = writer.sheets['Permissions']

                # Border & formatting
                thin = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                bold = Font(bold=True)

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                        if cell.row == 1:
                            cell.font = bold

                        if cell.column_letter == get_column_letter(df_excel.columns.get_loc('Needs Extension ? [y/n]') + 1):
                            cell.fill = yellow

                # Widths
                for col_idx, col in enumerate(df_excel.columns, 1):
                    max_length = max(df_excel[col].astype(str).map(len).max(), len(col))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

            excel_bytes = output_excel.getvalue()

            st.download_button(
                label="📘 Download Excel – clean format",
                data=excel_bytes,
                file_name=f"PRIAM - Special Permissions - {datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # -------------------------
            # --- EML FILE EXPORT ---
            # -------------------------
            table_html = df_to_html_table(df)

            email_body = f"""<html><body>
Dear manager,<br><br>
Could you please review below Special Permission about to expire for some of your team members and let us know if they need extension or not,<br><br>
{table_html}<br><br>
Thank you,<br>
Kr,<br><br>
Your IOC/BV
</body></html>
"""

            eml_content = (
                f"Subject: Special Permissions to be reviewed for your team"
                f"Content-Type: text/html; charset=UTF-8\n"
                f"\n"
                f"{email_body}"
            )

            st.download_button(
                label="📩 Download .eml email",
                data=eml_content.encode("utf-8"),
                file_name=f"PRIAM - Expiring Special Permissions - {datetime.now().strftime('%Y-%m-%d')}.eml",
                mime="message/rfc822"
            )

            st.success(f"✅ Extraction complete! {len(data)} email(s) processed.")

else:
    st.info("👈 Upload your .msg files using the sidebar to get started.")
    st.subheader("📖 Instructions")
    st.markdown("""
    1. Upload one or more `.msg` email files in the sidebar.  
    2. Click **"🚀 EXTRACT DATA"**.  
    3. Download Excel and .eml files.  
    4. Review the "Needs Extension ? [y/n]" column if needed.
    """)
